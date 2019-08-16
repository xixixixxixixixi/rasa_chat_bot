# -*- coding: utf-8 -*-

import json
import pandas as pd
import re
import openpyxl
import numpy as np
from matplotlib import pyplot as plt

#json_str=[{'date': '2019-07-10', 'open': 42.22, 'close': 40.81, 'high': 42.23, 'low': 40.75, 'volume': 886396, 'uOpen': 42.22, 'uClose': 40.81, 'uHigh': 42.23, 'uLow': 40.75, 'uVolume': 886396, 'change': 0,
# 'changePercent': 0, 'label': 'Jul 10', 'changeOverTime': 0}, {'date': '2019-07-11', 'open': 41.14, 'close': 40.85, 'high': 41.31, 'low': 40.39, 'volume': 715597, 'uOpen': 41.14, 'uClose': 40.85, 'uHigh': 41.31, 'uLow': 40.39, 'uVolume': 715597, 'change': 0.04, 'changePercent': 0.098, 'label': 'Jul 11', 'changeOverTime': 0.00098}, {'date': '2019-07-12', 'open': 40.83, 'close': 41.13, 'high': 41.57, 'low': 40.44, 'volume': 485854, 'uOpen': 40.83, 'uClose': 41.13, 'uHigh': 41.57, 'uLow': 40.44, 'uVolume': 485854, 'change': 0.28, 'changePercent': 0.6854, 'label': 'Jul 12', 'changeOverTime': 0.007841}, {'date': '2019-07-15', 'open': 41.61, 'close': 42.29, 'high': 42.6, 'low': 41.33, 'volume': 872170, 'uOpen': 41.61, 'uClose': 42.29, 'uHigh': 42.6, 'uLow': 41.33, 'uVolume': 872170, 'change': 1.16, 'changePercent': 2.8203, 'label': 'Jul 15', 'changeOverTime': 0.036266}, {'date': '2019-07-16', 'open': 42.03, 'close': 42.08, 'high': 42.75, 'low': 41.77, 'volume': 519835, 'uOpen': 42.03, 'uClose': 42.08, 'uHigh': 42.75, 'uLow': 41.77, 'uVolume': 519835, 'change': -0.21, 'changePercent': -0.4966, 'label': 'Jul 16', 'changeOverTime': 0.03112}, {'date': '2019-07-17', 'open': 42.08, 'close': 40.95, 'high': 42.14, 'low': 40.92, 'volume': 260624, 'uOpen': 42.08, 'uClose': 40.95, 'uHigh': 42.14, 'uLow': 40.92, 'uVolume': 260624, 'change': -1.13, 'changePercent': -2.6854, 'label': 'Jul 17', 'changeOverTime': 0.003431}, {'date': '2019-07-18', 'open': 41.01, 'close': 40.07, 'high': 41.01, 'low': 39.88, 'volume': 613783, 'uOpen': 41.01, 'uClose': 40.07, 'uHigh': 41.01, 'uLow': 39.88, 'uVolume': 613783, 'change': -0.88, 'changePercent': -2.149, 'label': 'Jul 18', 'changeOverTime': -0.018133}, {'date': '2019-07-19', 'open': 40.53, 'close': 40.44, 'high': 41.02, 'low': 40.42, 'volume': 662329, 'uOpen': 40.53, 'uClose': 40.44, 'uHigh': 41.02, 'uLow': 40.42, 'uVolume': 662329, 'change': 0.37, 'changePercent': 0.9234, 'label': 'Jul 19', 'changeOverTime': -0.009066}, {'date': '2019-07-22', 'open': 40.44, 'close': 39.88, 'high': 40.69, 'low': 39.56, 'volume': 902936, 'uOpen': 40.44, 'uClose': 39.88, 'uHigh': 40.69, 'uLow': 39.56, 'uVolume': 902936, 'change': -0.56, 'changePercent': -1.3848, 'label': 'Jul 22', 'changeOverTime': -0.022789}, {'date': '2019-07-23', 'open': 40.19, 'close': 40.61, 'high': 41.3, 'low': 40.19, 'volume': 1013273, 'uOpen': 40.19, 'uClose': 40.61, 'uHigh': 41.3, 'uLow': 40.19, 'uVolume': 1013273, 'change': 0.73, 'changePercent': 1.8305, 'label': 'Jul 23', 'changeOverTime': -0.004901}, {'date': '2019-07-24', 'open': 40.75, 'close': 41.02, 'high': 41.04, 'low': 40.42, 'volume': 274060, 'uOpen': 40.75, 'uClose': 41.02, 'uHigh': 41.04, 'uLow': 40.42, 'uVolume': 274060, 'change': 0.41, 'changePercent': 1.0096, 'label': 'Jul 24', 'changeOverTime': 0.005146}, {'date': '2019-07-25', 'open': 41.33, 'close': 40.37, 'high': 41.33, 'low': 40.04, 'volume': 454776, 'uOpen': 41.33, 'uClose': 40.37, 'uHigh': 41.33, 'uLow': 40.04, 'uVolume': 454776, 'change': -0.65, 'changePercent': -1.5846, 'label': 'Jul 25', 'changeOverTime': -0.010782}, {'date': '2019-07-26', 'open': 40.65, 'close': 40.33, 'high': 40.75, 'low': 40.03, 'volume': 460592, 'uOpen': 40.65, 'uClose': 40.33, 'uHigh': 40.75, 'uLow': 40.03, 'uVolume': 460592, 'change': -0.04, 'changePercent': -0.0991, 'label': 'Jul 26', 'changeOverTime': -0.011762}, {'date': '2019-07-29', 'open': 40.29, 'close': 40.36, 'high': 40.73, 'low': 40.06, 'volume': 619979, 'uOpen': 40.29, 'uClose': 40.36, 'uHigh': 40.73, 'uLow': 40.06, 'uVolume': 619979, 'change': 0.03, 'changePercent': 0.0744, 'label': 'Jul 29', 'changeOverTime': -0.011027}, {'date': '2019-07-30', 'open': 39.9, 'close': 40.08, 'high': 40.35, 'low': 39.47, 'volume': 492151, 'uOpen': 39.9, 'uClose': 40.08, 'uHigh': 40.35, 'uLow': 39.47, 'uVolume': 492151, 'change': -0.28, 'changePercent': -0.6938, 'label': 'Jul 30', 'changeOverTime': -0.017888}, {'date': '2019-07-31', 'open': 40.05, 'close': 39.12, 'high': 40.05, 'low': 38.63, 'volume': 865797, 'uOpen': 40.05, 'uClose': 39.12, 'uHigh': 40.05, 'uLow': 38.63, 'uVolume': 865797, 'change': -0.96, 'changePercent': -2.3952, 'label': 'Jul 31', 'changeOverTime': -0.041411}, {'date': '2019-08-01', 'open': 39.15, 'close': 37.36, 'high': 39.51, 'low': 37.11, 'volume': 1064391, 'uOpen': 39.15, 'uClose': 37.36, 'uHigh': 39.51, 'uLow': 37.11, 'uVolume': 1064391, 'change': -1.76, 'changePercent': -4.499, 'label': 'Aug 1', 'changeOverTime': -0.084538}, {'date': '2019-08-02', 'open': 37.07, 'close': 37.3, 'high': 37.77, 'low': 36.57, 'volume': 629357, 'uOpen': 37.07, 'uClose': 37.3, 'uHigh': 37.77, 'uLow': 36.57, 'uVolume': 629357, 'change': -0.06, 'changePercent': -0.1606, 'label': 'Aug 2', 'changeOverTime': -0.086008}, {'date': '2019-08-05', 'open': 35.76, 'close': 34.41, 'high': 35.76, 'low': 33.95, 'volume': 1208231, 'uOpen': 35.76, 'uClose': 34.41, 'uHigh': 35.76, 'uLow': 33.95, 'uVolume': 1208231, 'change': -2.89, 'changePercent': -7.748, 'label': 'Aug 5', 'changeOverTime': -0.156824}, {'date': '2019-08-06', 'open': 35.09, 'close': 35.23, 'high': 36.04, 'low': 35, 'volume': 1201426, 'uOpen': 35.09, 'uClose': 35.23, 'uHigh': 36.04, 'uLow': 35, 'uVolume': 1201426, 'change': 0.82, 'changePercent': 2.383, 'label': 'Aug 6', 'changeOverTime': -0.136731}, {'date': '2019-08-07', 'open': 34.66, 'close': 35.21, 'high': 35.6, 'low': 34.4, 'volume': 1013158, 'uOpen': 34.66, 'uClose': 35.21, 'uHigh': 35.6, 'uLow': 34.4, 'uVolume': 1013158, 'change': -0.02, 'changePercent': -0.0568, 'label': 'Aug 7', 'changeOverTime': -0.137221}, {'date': '2019-08-08', 'open': 35.59, 'close': 35.71, 'high': 36, 'low': 35.28, 'volume': 508580, 'uOpen': 35.59, 'uClose': 35.71, 'uHigh': 36, 'uLow': 35.28, 'uVolume': 508580, 'change': 0.5, 'changePercent': 1.4201, 'label': 'Aug 8', 'changeOverTime': -0.124969}, {'date': '2019-08-09', 'open': 35.34, 'close': 34.52, 'high': 35.4, 'low': 34.52, 'volume': 1149580, 'uOpen': 35.34, 'uClose': 34.52, 'uHigh': 35.4, 'uLow': 34.52, 'uVolume': 1149580, 'change': -1.19, 'changePercent': -3.3324, 'label': 'Aug 9', 'changeOverTime': -0.154129}]

def text2pic(ID,json_str):
    y=[]
    x=np.arange(0,len(json_str))
    filename='res.jpg'
    for list_mem in json_str:
        #print(list_mem[ID])
        #print(type(list_mem[ID]))
        y.append(list_mem[ID])
    firstdate=json_str[0]["date"]
    plt.title("stock trends")
    plt.xlabel("0 means "+firstdate)
    plt.ylabel(ID)
    print(x)
    print(y)
    plt.plot(x, y)
    plt.savefig(filename)
    return filename

#text2pic("open",json_str)


def text2sheet(json_str):
    with open('temp.txt','w', encoding = 'UTF-8') as fr:
         for list_mem in json_str:
             list_mem=str(list_mem)
             list_mem=list_mem.replace("'", '"')
             fr.write('['+list_mem+']'+"\n")
    
    data = [] # 用于存储每一行的Json数据
    with open('temp.txt','r', encoding = 'UTF-8') as fr:
        for line in fr:
            j = json.loads(line)
            data.append(j)
    
    df = pd.DataFrame() # 最后转换得到的结果
    for line in data:
        for i in line:
            df1 = pd.DataFrame([i])
            df = df.append(df1)
    
    # 在excel表格的第1列写入, 不写入index
    df.to_excel('data.xlsx', sheet_name='Data', startcol=0, index=False)
    
def checksheet(date,item):
    wb=openpyxl.load_workbook('data.xlsx')
    ws=wb.active
    x=''
    y=''
    flag1,flag2=0,0
    #sheet=wb.get_sheet_by_name('Data')
    for i in range(ord("A"),ord("O")+1):
        if(ws[chr(i)+'1'].value==item):
            x=chr(i)
            flag1=1
            break
    for j in range(1,22):
        if(ws['E'+str(j)].value==date):
            y=str(j)
            flag2=1
            break
    if(flag1==1 and flag2==1):
        return ws[x+y].value
    else:
        return None
    
def date_extract(message):
    x = re.compile(r'\d{4}-\d{2}-\d{2}')
    date=x.findall(message)
    return date
    
            
    
    
